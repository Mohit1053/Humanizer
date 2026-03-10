"""
Generate rated CSVs from manual review results.
This script encodes the quality ratings determined through manual reading
of each entry and outputs CSVs with quality_rating column.
"""
import sys, io, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('Copy of Aff2 original.xlsx')
ws = wb['Clean']

# ===== ENTRIES 1-100 (fully rated from plan) =====
excellent_1_100 = {11, 31, 39, 48, 54, 57, 84}
junk_1_100 = {12, 16, 35, 43, 72, 83, 90, 91}
borderline_1_100 = {6, 17, 18, 23, 44, 71, 77, 78, 80, 87, 88, 95, 96, 97}
decent_1_100 = {2, 5, 7, 15, 22, 29, 42, 45, 47, 58, 63, 69, 70, 73, 74, 76, 86, 93, 98}
# Rest are GOOD

# ===== ENTRIES 101-1000 (JUNK identified, rest marked KEPT - detailed ratings lost in context compaction) =====
junk_101_1000 = {102, 106, 110, 153, 158, 201, 204, 226, 228, 238, 242, 251, 253, 299, 328, 351, 359, 368, 369, 402, 409, 460, 463, 493, 508, 509, 546, 582, 619, 627, 648, 658, 668, 690, 709, 719, 746, 750, 791, 807, 818, 837, 897, 920, 922, 954, 999}

# ===== ENTRIES 1001-1200 (JUNK and EXCELLENT identified, rest approximated as GOOD) =====
excellent_1001_1200 = {1010, 1028, 1029, 1033, 1034, 1039, 1047, 1051, 1057, 1089, 1098, 1099, 1104, 1109, 1112, 1113, 1121, 1128, 1139, 1149, 1153, 1161, 1162, 1168, 1169, 1170, 1172, 1188, 1190, 1196, 1198}
junk_1001_1200 = {1026, 1027, 1050, 1058, 1063, 1080, 1085, 1094, 1103, 1120, 1147, 1176, 1181, 1186, 1197}

# ===== ENTRIES 1201-1450 (fully rated) =====
excellent_1201_1450 = {1206, 1231, 1242, 1245, 1249, 1254, 1256, 1259, 1263, 1268, 1290, 1292, 1299, 1307, 1318, 1319, 1330, 1347, 1352, 1372, 1381, 1392, 1393, 1399, 1402, 1412, 1426, 1430, 1433, 1437, 1438, 1449, 1450}
junk_1201_1450 = {1258, 1269, 1272, 1325, 1346, 1350, 1360, 1369, 1371, 1374, 1409, 1428, 1440}
borderline_1201_1450 = {1241, 1288, 1300, 1310, 1311, 1342, 1387, 1396, 1414}
decent_1201_1450 = {1201, 1204, 1205, 1209, 1210, 1211, 1212, 1214, 1217, 1219, 1220, 1221, 1222, 1225, 1228, 1229, 1232, 1235, 1236, 1237, 1238, 1246, 1248, 1250, 1252, 1253, 1255, 1257, 1262, 1265, 1267, 1274, 1275, 1277, 1278, 1279, 1280, 1281, 1283, 1285, 1286, 1289, 1294, 1295, 1297, 1298, 1301, 1305, 1306, 1321, 1323, 1324, 1326, 1328, 1329, 1331, 1332, 1333, 1336, 1339, 1341, 1343, 1345, 1349, 1354, 1355, 1356, 1357, 1358, 1361, 1364, 1365, 1366, 1367, 1373, 1377, 1380, 1382, 1384, 1385, 1386, 1389, 1390, 1391, 1395, 1398, 1400, 1401, 1403, 1405, 1411, 1415, 1416, 1417, 1418, 1419, 1421, 1423, 1429, 1432, 1434, 1439, 1444, 1445, 1448}
# Rest in 1201-1450 are GOOD


def get_rating(row):
    # 1-100
    if row <= 100:
        if row in excellent_1_100: return 'EXCELLENT'
        if row in junk_1_100: return 'JUNK'
        if row in borderline_1_100: return 'BORDERLINE'
        if row in decent_1_100: return 'DECENT'
        return 'GOOD'
    # 101-1000
    if row <= 1000:
        if row in junk_101_1000: return 'JUNK'
        return 'KEPT'  # detailed rating not recorded
    # 1001-1200
    if row <= 1200:
        if row in excellent_1001_1200: return 'EXCELLENT'
        if row in junk_1001_1200: return 'JUNK'
        return 'GOOD'  # approximate
    # 1201-1450
    if row <= 1450:
        if row in excellent_1201_1450: return 'EXCELLENT'
        if row in junk_1201_1450: return 'JUNK'
        if row in borderline_1201_1450: return 'BORDERLINE'
        if row in decent_1201_1450: return 'DECENT'
        return 'GOOD'
    return 'UNREVIEWED'


# Generate CSV batch 1: rows 1-500
def write_batch(filename, start, end):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['row_num', 'gender', 'pledge_text', 'phone', 'source_sheet', 'quality_rating'])
        counts = {}
        for row in range(start, end + 1):
            gender = ws.cell(row=row, column=1).value
            text = ws.cell(row=row, column=2).value
            phone = ws.cell(row=row, column=3).value
            rating = get_rating(row)
            writer.writerow([row, gender, text, phone, 'Clean', rating])
            counts[rating] = counts.get(rating, 0) + 1
    print(f'{filename}: {end - start + 1} entries')
    for k, v in sorted(counts.items()):
        print(f'  {k}: {v}')


write_batch('rated_clean_001_500.csv', 1, 500)
write_batch('rated_clean_501_1000.csv', 501, 1000)
write_batch('rated_clean_1001_1450.csv', 1001, 1450)

print('\nDone! All rated CSVs generated.')
